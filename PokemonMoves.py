import requests
import concurrent.futures
import openpyxl


def obtener_nombre_movimiento(move_id):
    url = f"https://pokeapi.co/api/v2/move/{move_id}"
    response = requests.get(url)
    if response.status_code == 200:
        data = response.json()
        return data['name']
    else:
        return None

def obtener_nombres_movimientos(start_id, end_id):
    nombres_movimientos = []
    for move_id in range(start_id, end_id + 1):
        nombre_movimiento = obtener_nombre_movimiento(move_id)
        if nombre_movimiento:
            nombres_movimientos.append(nombre_movimiento)
    return nombres_movimientos

def imprimir_salto_linea(moves):
    for move in moves:
        print(move)
    print('------------')

def imprimir_nombres_movimientos():
    start_id = 1
    end_id = 900
    step = 20
    num_procesos = 10

    with concurrent.futures.ThreadPoolExecutor(max_workers=num_procesos) as executor:
        futures = [executor.submit(obtener_nombres_movimientos, i, i + step - 1) for i in range(start_id, end_id + 1, step)]

        nombres_totales = []
        for future in concurrent.futures.as_completed(futures):
            nombres_movimientos = future.result()
            imprimir_salto_linea(nombres_movimientos)
            nombres_totales.extend(nombres_movimientos)
        guardar_movimientos(nombres_totales)

def guardar_movimientos(moveset):
    wb = openpyxl.Workbook()
    sheet1 = wb.active

    for i in range(len(moveset)):
        sheet1.cell(i+1, 1, moveset[i])

    sheet1.insert_rows(idx=1)
    sheet1.cell(1,1,'Movimientos')
    wb.save('Moveset.xlsx')

if __name__ == "__main__":
    imprimir_nombres_movimientos()
